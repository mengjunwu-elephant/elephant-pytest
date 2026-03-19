# -*- coding: utf-8 -*-
from __future__ import annotations

import os
from typing import Any, Optional, Sequence

from openpyxl import load_workbook


def get_test_data_from_excel(
    file: str,
    sheet_name: str,
    required_columns: Optional[Sequence[str]] = None,
) -> list[dict[str, Any]]:
    """
    读取 Excel 首行为字段名，自第 2 行起每行一条用例字典。

    :param file: xlsx 路径
    :param sheet_name: 工作表名（与代码中 get_test_data_from_excel(..., sheet_name) 一致）
    :param required_columns: 若给出，校验首行必须包含这些列名
    :return: 字典列表；全空行会被跳过
    :raises FileNotFoundError: 文件不存在
    :raises KeyError: 工作表不存在
    :raises ValueError: 首行列名非法或缺少 required_columns
    """
    if not os.path.isfile(file):
        raise FileNotFoundError(file)

    wb = load_workbook(file, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"工作表 {sheet_name!r} 不存在，当前工作簿包含: {wb.sheetnames!r}"
            )
        sh = wb[sheet_name]
        row = sh.max_row
        column = sh.max_column
        data: list[dict[str, Any]] = []
        keys: list[Any] = []
        for i in range(1, column + 1):
            keys.append(sh.cell(1, i).value)
        if any(k is None or (isinstance(k, str) and k.strip() == "") for k in keys):
            raise ValueError("Excel 首行存在空列名，请删除空列或填写表头")

        str_keys = [str(k).strip() for k in keys]

        if required_columns is not None:
            missing = set(required_columns) - set(str_keys)
            if missing:
                raise ValueError(f"Excel 缺少必填列: {sorted(missing)}")

        for i in range(2, row + 1):
            temp: dict[str, Any] = {}
            for j in range(1, column + 1):
                temp[str_keys[j - 1]] = sh.cell(i, j).value
            if all(
                v is None or (isinstance(v, str) and v.strip() == "")
                for v in temp.values()
            ):
                continue
            data.append(temp)
        return data
    finally:
        wb.close()


if __name__ == "__main__":
    get_test_data_from_excel(r"../test_data/mercury.xlsx", "Sheet1")
