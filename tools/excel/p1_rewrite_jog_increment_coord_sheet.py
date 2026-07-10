# -*- coding: utf-8 -*-
"""重写 jog_increment_coord sheet：初始点位 + 四轴 ±20mm + 异常边界。"""
from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")
SHEET = "jog_increment_coord"

MIN_COORDS = [-350, -362.43, -186.265, -180.0]
MAX_COORDS = [360.43, 362.43, 93.44, 180.0]
AXIS_NAMES = ["X", "Y", "Z", "Rx"]
STEP = 20
SPEED = 50


def _oor_increment(axis_idx: int) -> int:
    return int(abs(MIN_COORDS[axis_idx]) + abs(MAX_COORDS[axis_idx])) + 1


def _build_rows() -> tuple[list[str], list[list]]:
    headers = ["ID", "title", "api", "axis", "increment", "speed", "expect_data", "test_type"]
    rows: list[list] = []
    rid = 1
    for i, name in enumerate(AXIS_NAMES):
        axis = i + 1
        for inc, label in ((STEP, "正向"), (-STEP, "负向")):
            rows.append(
                [
                    rid,
                    f"初始点位_{name}轴步进{label}{abs(inc)}mm",
                    "jog_increment_coord",
                    axis,
                    inc,
                    SPEED,
                    "ok",
                    "normal",
                ]
            )
            rid += 1
    rows.extend(
        [
            [rid, "设置速度超限0", "jog_increment_coord", 1, STEP, 0, None, "exception"],
            [rid + 1, "设置速度超限101", "jog_increment_coord", 1, STEP, 101, None, "exception"],
            [rid + 2, "设置坐标轴号超限0", "jog_increment_coord", 0, STEP, SPEED, None, "exception"],
            [rid + 3, "设置坐标轴号超限5", "jog_increment_coord", 5, STEP, SPEED, None, "exception"],
        ]
    )
    rid += 4
    for i, name in enumerate(AXIS_NAMES):
        axis = i + 1
        oor = _oor_increment(i)
        rows.append(
            [rid, f"{name}轴增量正向超限", "jog_increment_coord", axis, oor, SPEED, None, "exception"]
        )
        rid += 1
        rows.append(
            [rid, f"{name}轴增量负向超限", "jog_increment_coord", axis, -oor, SPEED, None, "exception"]
        )
        rid += 1
    return headers, rows


def main() -> None:
    shutil.copy2(XLSX, XLSX + ".bak_jog_increment_coord")
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    headers, rows = _build_rows()
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(XLSX)
    print(f"Rewrote {SHEET}: {len(rows)} rows")


if __name__ == "__main__":
    main()
