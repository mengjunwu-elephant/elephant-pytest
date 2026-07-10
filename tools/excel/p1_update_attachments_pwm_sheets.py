# -*- coding: utf-8 -*-
"""Attachments 表：移除 open/close_laser，添加 PWM 五接口 sheet（最小用例集）。"""
from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "test_data", "UltraArm_P1_Attachments.xlsx")

REMOVE = ("open_laser", "close_laser")

SHEETS: dict[str, tuple[list[str], list[list]]] = {
    "quick_off_laser": (
        ["ID", "title", "api", "state", "expect_data", "test_type"],
        [
            [1, "打开激光PWM模式", "quick_off_laser", 1, "ok", "normal"],
            [2, "关闭激光PWM模式", "quick_off_laser", 0, "ok", "normal_off"],
            [3, "激光模式状态越界", "quick_off_laser", 2, None, "exception"],
        ],
    ),
    "set_pwm_laser": (
        ["ID", "title", "api", "p_value", "expect_data", "test_type"],
        [
            [1, "设置激光PWM最小档位", "set_pwm_laser", 0, "ok", "normal"],
            [2, "设置激光PWM最大档位", "set_pwm_laser", 255, "ok", "normal"],
            [3, "激光PWM档位越界", "set_pwm_laser", 256, None, "exception"],
        ],
    ),
    "quick_off_custom_pwm": (
        ["ID", "title", "api", "state", "expect_data", "test_type"],
        [
            [1, "打开自定义PWM模式", "quick_off_custom_pwm", 1, "ok", "normal"],
            [2, "关闭自定义PWM模式", "quick_off_custom_pwm", 0, "ok", "normal_off"],
            [3, "自定义PWM模式状态越界", "quick_off_custom_pwm", 2, None, "exception"],
        ],
    ),
    "set_pwm_custom": (
        ["ID", "title", "api", "p_value", "expect_data", "test_type"],
        [
            [1, "设置自定义PWM最小档位", "set_pwm_custom", 0, "ok", "normal"],
            [2, "设置自定义PWM最大档位", "set_pwm_custom", 255, "ok", "normal"],
            [3, "自定义PWM档位越界", "set_pwm_custom", 256, None, "exception"],
        ],
    ),
    "get_pwm_status": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取PWM默认关闭状态", "get_pwm_status", "[0,0,0,0]", "normal"]],
    ),
}


def _apply_pwm_sheets(wb_path: str) -> None:
    if not os.path.isfile(wb_path):
        raise FileNotFoundError(wb_path)
    tmp = wb_path + ".tmp_write"
    shutil.copy2(wb_path, wb_path + ".bak_pwm")
    wb = load_workbook(wb_path)
    for name in REMOVE:
        if name in wb.sheetnames:
            del wb[name]
    for name, (headers, rows) in SHEETS.items():
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(tmp)
    wb.close()
    os.replace(tmp, wb_path)


def main() -> None:
    try:
        _apply_pwm_sheets(XLSX)
        print(f"Updated PWM sheets: {XLSX}")
    except PermissionError:
        fallback = XLSX.replace(".xlsx", "_pwm_updated.xlsx")
        if os.path.isfile(fallback):
            _apply_pwm_sheets(fallback)
        else:
            shutil.copy2(XLSX, fallback)
            _apply_pwm_sheets(fallback)
        print(f"主文件被占用，已写入 {fallback}，请关闭 Excel 后重新运行本脚本。")
    print("Removed open_laser/close_laser; PWM 五接口 sheet 已写入（最小用例集）。")


if __name__ == "__main__":
    main()
