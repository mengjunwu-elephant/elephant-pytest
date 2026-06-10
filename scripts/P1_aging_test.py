import argparse
import os
import signal
import threading
import time
from typing import Any

from openpyxl import Workbook
from pymycobot import UltraArmP1

from common1 import logger


# P1 joint limits from your requirement.
# J1: +/-158, J2: -18~85, J3: 89~190, J4: +/-179
INIT_ANGLES = [0.0, 10.0, 110.0, 0.0]
# 笛卡尔运动前回到该姿态，减轻 J2/J3 耦合（LimitError:6）；与接口测试 coords_init 一致
COORDS_INIT_ANGLES = [0.0, 10.0, 110.0, 0.0]

# [X,Y,Z,Rx] 示教空间软限位（与现场文档一致时可再改）
COORD_LIMITS = (
    (-301.7, 360.5),
    (-360.5, 360.5),
    (-157.0, 91.0),
    (-180.0, 180.0),
)

JOINT_CASES = {
    "J1": {"min": [-158.0, 10.0, 110.0, 0.0], "max": [158.0, 10.0, 110.0, 0.0]},
    "J2": {"min": [0.0, -18.0, 120.0, 0.0], "max": [0.0, 85.0, 110.0, 0.0]},
    "J3": {"min": [0.0, 10.0, 89.0, 0.0], "max": [150, 60.0, 190.0, 0.0]},
    "J4": {"min": [0.0, 10.0, 110.0, -179.0], "max": [0.0, 10.0, 110.0, 179.0]},
}

ANGLE_TOLERANCE_DEG = 0.5
# 坐标：XYZ 允许误差 1 mm；Rx 为角度分量，允许误差 1°
COORD_TOLERANCE_MM = 1.0
COORD_TOLERANCE_RX_DEG = 1.0

MAX_CONSECUTIVE_MONITOR_INVALID = 20

AXIS_NAMES = ("X", "Y", "Z", "Rx")


stop_threads = threading.Event()
stats_lock = threading.Lock()

# 仅用于 finally 里 close；切勿在信号处理器里调用 mc.stop()（易与 get_run_status 等争用串口锁卡死）
_mc_holder: dict[str, Any] = {"mc": None}


def _request_shutdown(_signum: int | None = None, _frame: Any = None) -> None:
    """Ctrl+C：只置位标志。stop() 放在 wait_motion_done 的同线程循环里执行，避免死锁。"""
    if stop_threads.is_set():
        return
    logger.info("收到停止请求（Ctrl+C 或 SIGINT/SIGBREAK），正在停止测试…")
    stop_threads.set()


def _install_signal_handlers() -> None:
    """改善 Windows 下 Ctrl+C：避免无限 join 阻塞导致无法响应。"""
    try:
        signal.signal(signal.SIGINT, _request_shutdown)
    except (ValueError, OSError) as e:
        logger.warning("注册 SIGINT 失败：%s", e)
    # Windows：Ctrl+Break
    if hasattr(signal, "SIGBREAK"):
        try:
            signal.signal(signal.SIGBREAK, _request_shutdown)
        except (ValueError, OSError):
            pass


def _join_threads_until_done(
    t_move: threading.Thread,
    t_monitor: threading.Thread,
    poll: float = 0.5,
    max_wait_sec: float | None = None,
) -> None:
    """带超时的 join，让主线程能处理 KeyboardInterrupt / 信号；可选总等待上限。"""
    deadline = time.time() + max_wait_sec if max_wait_sec is not None else None
    while t_move.is_alive() or t_monitor.is_alive():
        if deadline is not None and time.time() > deadline:
            logger.warning(
                "等待子线程结束已超时（%.0f 秒），将继续执行清理；必要时请结束 Python 进程",
                max_wait_sec or 0,
            )
            break
        t_move.join(timeout=poll)
        t_monitor.join(timeout=poll)


def wait_motion_done(mc: UltraArmP1, timeout_sec: float = 30.0) -> bool:
    """正常停稳返回 True；超时或用户中断返回 False。

    用户按 Ctrl+C 时仅在信号里置 stop_threads；此处（同一线程）再调用 mc.stop()，
    避免在信号处理器内 stop 与 get_run_status 嵌套争用 pymycobot 串口锁导致进程卡死。
    """
    start = time.time()
    time.sleep(0.2)
    user_stop_dispatched = False
    while True:
        if time.time() - start > timeout_sec:
            logger.warning("等待运动停止超时：%.1f 秒", timeout_sec)
            return False

        if stop_threads.is_set() and not user_stop_dispatched:
            user_stop_dispatched = True
            try:
                mc.stop()
                logger.info("用户停止：已在等待循环内调用 mc.stop()")
            except Exception as e:
                logger.warning("mc.stop() 异常：%s", e)

        try:
            moving = mc.get_run_status()
        except Exception as e:
            logger.warning("get_run_status 异常：%s", e)
            time.sleep(0.05)
            continue

        if not moving:
            time.sleep(0.2)
            return not stop_threads.is_set()

        time.sleep(0.05)


def angles_reached(target: list[float], actual: list[float], tol: float = ANGLE_TOLERANCE_DEG) -> bool:
    if len(target) != len(actual):
        return False
    for t, a in zip(target, actual):
        if abs(float(a) - float(t)) > tol:
            return False
    return True


def move_to_coords_init(mc: UltraArmP1, speed: int) -> None:
    mc.set_angles(list(COORDS_INIT_ANGLES), speed)
    wait_motion_done(mc, timeout_sec=30.0)


def safe_get_coords(mc: UltraArmP1) -> list[float] | None:
    try:
        data = mc.get_coords_info()
        if data in (-1, None):
            return None
        return [float(x) for x in data]
    except Exception as e:
        logger.warning("读取 get_coords_info 异常：%s", e)
        return None


def coords_reached(
    target: list[float],
    actual: list[float],
    tol_mm: float = COORD_TOLERANCE_MM,
    tol_rx: float = COORD_TOLERANCE_RX_DEG,
) -> bool:
    if len(target) != len(actual) or len(target) != 4:
        return False
    for i in range(3):
        if abs(float(actual[i]) - float(target[i])) > tol_mm:
            return False
    if abs(float(actual[3]) - float(target[3])) > tol_rx:
        return False
    return True


def symmetric_coord_targets(
    base: list[float], axis_index: int, delta_mm_or_deg: float
) -> tuple[list[float] | None, list[float] | None]:
    """在限位内生成该轴对称负向/正向目标；步长不足则返回 (None, None)。"""
    lo, hi = COORD_LIMITS[axis_index]
    b = float(base[axis_index])
    step = float(delta_mm_or_deg)
    max_neg = b - lo
    max_pos = hi - b
    step = min(step, max_neg, max_pos)
    if step < 0.5:
        return None, None
    neg = list(base)
    pos = list(base)
    neg[axis_index] = b - step
    pos[axis_index] = b + step
    return neg, pos


def safe_get_angles(mc: UltraArmP1) -> list[float] | None:
    try:
        data = mc.get_angles_info()
        if data in (-1, None):
            return None
        return [float(x) for x in data]
    except Exception as e:
        logger.warning("读取 get_angles_info 异常：%s", e)
        return None


def move_to_limit(
    mc: UltraArmP1,
    joint_name: str,
    target: list[float],
    direction: str,
    speed: int,
    movement_stats: dict[str, int],
) -> tuple[Any, int]:
    """
    Move to target limit and return:
    - movement_time: float seconds | string reason
    - fail_flag: 0 success, 1 failure
    """
    if stop_threads.is_set():
        return "stopped", 1

    with stats_lock:
        movement_stats["angles_attempts"] += 1

    start = time.time()
    try:
        mc.set_angles(target, speed)
    except Exception as e:
        with stats_lock:
            movement_stats["angles_failed"] += 1
        logger.warning("%s %s set_angles 调用异常：%s", joint_name, direction, e)
        return f"set_error: {e}", 1

    if not wait_motion_done(mc, timeout_sec=20.0):
        with stats_lock:
            movement_stats["angles_failed"] += 1
        return "timeout_or_stop", 1

    actual = safe_get_angles(mc)
    if actual is None:
        with stats_lock:
            movement_stats["angles_failed"] += 1
        return "angles_read_failed", 1

    if not angles_reached(target, actual):
        with stats_lock:
            movement_stats["angles_failed"] += 1
        logger.warning(
            "%s %s 未到位，目标=%s，实际=%s",
            joint_name,
            direction,
            target,
            actual,
        )
        return f"not_reached:{actual}", 1

    movement_time = round(time.time() - start, 3)
    logger.info("%s %s 完成，用时 %s 秒", joint_name, direction, movement_time)
    return movement_time, 0


def move_one_coord(
    mc: UltraArmP1,
    target: list[float],
    speed: int,
    movement_stats: dict[str, int],
) -> tuple[Any, int, str]:
    """
    set_coords 后校验笛卡尔到位：XYZ ≤COORD_TOLERANCE_MM，Rx ≤COORD_TOLERANCE_RX_DEG。
    关节指令未给定，不对关节角做数值比对；关节测试阶段单独按 ±ANGLE_TOLERANCE_DEG 校验。
    """
    if stop_threads.is_set():
        return "stopped", 1, ""

    with stats_lock:
        movement_stats["coords_attempts"] += 1

    move_to_coords_init(mc, speed)
    start = time.time()
    try:
        mc.set_coords(target, speed)
    except Exception as e:
        with stats_lock:
            movement_stats["coords_failed"] += 1
        logger.warning("set_coords 调用异常：%s", e)
        return f"set_error:{e}", 1, ""

    if not wait_motion_done(mc, timeout_sec=25.0):
        with stats_lock:
            movement_stats["coords_failed"] += 1
        return "timeout_or_stop", 1, ""

    actual_coord = safe_get_coords(mc)
    if actual_coord is None:
        with stats_lock:
            movement_stats["coords_failed"] += 1
        return "coords_read_failed", 1, ""

    if not coords_reached(target, actual_coord):
        with stats_lock:
            movement_stats["coords_failed"] += 1
        logger.warning("坐标未到位，目标=%s，实际=%s", target, actual_coord)
        return f"coord_not_reached:{actual_coord}", 1, str(actual_coord)

    movement_time = round(time.time() - start, 3)
    feedback = str(actual_coord)
    logger.info(
        "坐标运动到位（XYZ≤%s mm，Rx≤%s°），用时 %s 秒，反馈=%s",
        COORD_TOLERANCE_MM,
        COORD_TOLERANCE_RX_DEG,
        movement_time,
        feedback,
    )
    return movement_time, 0, feedback


def coord_axis_cycle(
    mc: UltraArmP1,
    cycle_index: int,
    speed: int,
    delta_xyz_mm: float,
    delta_rx_deg: float,
    movement_stats: dict[str, int],
    ws_coords: Any,
    wb: Workbook,
    save_path: str,
) -> None:
    """先回到 coords_init，读取基准坐标，再对 X/Y/Z/Rx 分别做对称负向、正向运动并校验到位。"""
    move_to_coords_init(mc, speed)
    base = safe_get_coords(mc)
    if base is None:
        logger.warning("第 %s 轮：无法读取基准坐标，跳过本轮回笛卡尔运动", cycle_index)
        return

    for axis_idx, axis_name in enumerate(AXIS_NAMES):
        if stop_threads.is_set():
            break
        d = delta_xyz_mm if axis_idx < 3 else delta_rx_deg
        neg_t, pos_t = symmetric_coord_targets(base, axis_idx, d)
        if neg_t is None or pos_t is None:
            logger.warning(
                "第 %s 轮 %s 轴：当前基准下可动行程不足，跳过（基准[%s]=%s，限位=%s~%s）",
                cycle_index,
                axis_name,
                axis_name,
                base[axis_idx],
                COORD_LIMITS[axis_idx][0],
                COORD_LIMITS[axis_idx][1],
            )
            continue

        neg_time, neg_fail, neg_fb = move_one_coord(mc, neg_t, speed, movement_stats)
        if stop_threads.is_set():
            break
        pos_time, pos_fail, pos_fb = move_one_coord(mc, pos_t, speed, movement_stats)

        ws_coords.append(
            [
                cycle_index,
                axis_name,
                neg_time,
                pos_time,
                int(neg_fail),
                int(pos_fail),
                str(neg_t),
                str(pos_t),
                neg_fb,
                pos_fb,
                str(base),
            ]
        )
        try:
            wb.save(save_path)
        except Exception as e:
            logger.warning("保存工作簿告警：%s", e)


def movement_thread(
    mc: UltraArmP1,
    speed: int,
    ws_movement: Any,
    ws_coords: Any,
    wb: Workbook,
    save_path: str,
    movement_stats: dict[str, int],
    cycles: int,
    delta_xyz_mm: float,
    delta_rx_deg: float,
) -> None:
    cycle_index = 0
    while not stop_threads.is_set():
        cycle_index += 1
        logger.info("开始老化循环：第 %s 轮", cycle_index)

        for joint_name, limits in JOINT_CASES.items():
            if stop_threads.is_set():
                break

            neg_time, neg_fail = move_to_limit(
                mc,
                joint_name,
                limits["min"],
                "负向",
                speed,
                movement_stats,
            )
            if stop_threads.is_set():
                break

            pos_time, pos_fail = move_to_limit(
                mc,
                joint_name,
                limits["max"],
                "正向",
                speed,
                movement_stats,
            )

            ws_movement.append(
                [
                    cycle_index,
                    joint_name,
                    neg_time,
                    pos_time,
                    int(neg_fail),
                    int(pos_fail),
                    limits["min"].__str__(),
                    limits["max"].__str__(),
                ]
            )

            try:
                wb.save(save_path)
            except Exception as e:
                logger.warning("保存工作簿告警：%s", e)

        if not stop_threads.is_set():
            logger.info("第 %s 轮：开始笛卡尔 X/Y/Z/Rx 正反向运动", cycle_index)
            coord_axis_cycle(
                mc,
                cycle_index,
                speed,
                delta_xyz_mm,
                delta_rx_deg,
                movement_stats,
                ws_coords,
                wb,
                save_path,
            )

        with stats_lock:
            attempts = movement_stats["angles_attempts"]
            failed = movement_stats["angles_failed"]
            ca = movement_stats.get("coords_attempts", 0)
            cf = movement_stats.get("coords_failed", 0)
        logger.info(
            "当前运动统计：关节发送=%s，关节失败=%s；坐标发送=%s，坐标失败=%s",
            attempts,
            failed,
            ca,
            cf,
        )

        if cycles > 0 and cycle_index >= cycles:
            logger.info("达到设定循环次数=%s，准备停止测试", cycles)
            stop_threads.set()
            break

        time.sleep(0.1)


def monitor_thread(
    mc: UltraArmP1,
    poll_s: float,
    monitor_stats: dict[str, int],
    monitor_events: list[list[Any]],
) -> None:
    consecutive_invalid = 0

    while not stop_threads.is_set():
        ts = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        try:
            err_info = mc.get_error_information()
            run_status = mc.get_run_status()
            motor_enable = mc.get_motor_enable_status()
            with stats_lock:
                monitor_stats["poll_count"] += 1

            invalid_this_round = 0
            if err_info in (-1, None):
                invalid_this_round += 1
                with stats_lock:
                    monitor_stats["error_info_invalid"] += 1
            if run_status in (-1, None):
                invalid_this_round += 1
                with stats_lock:
                    monitor_stats["run_status_invalid"] += 1
            if motor_enable in (-1, None):
                invalid_this_round += 1
                with stats_lock:
                    monitor_stats["motor_enable_invalid"] += 1

            # Keep event rows only for abnormal situations.
            if invalid_this_round > 0 or (err_info not in (0, "0", "", [], None)):
                monitor_events.append(
                    [
                        ts,
                        str(err_info),
                        str(run_status),
                        str(motor_enable),
                        invalid_this_round,
                    ]
                )

            if invalid_this_round > 0:
                consecutive_invalid += 1
            else:
                consecutive_invalid = 0

            if consecutive_invalid >= MAX_CONSECUTIVE_MONITOR_INVALID:
                logger.error(
                    "监控连续异常已达到 %s 次，停止测试",
                    MAX_CONSECUTIVE_MONITOR_INVALID,
                )
                stop_threads.set()
                break

        except Exception as e:
            with stats_lock:
                monitor_stats["monitor_exceptions"] += 1
            consecutive_invalid += 1
            logger.warning("监控线程异常：%s", e)
            monitor_events.append([ts, f"exception:{e}", "", "", 1])
            if consecutive_invalid >= MAX_CONSECUTIVE_MONITOR_INVALID:
                stop_threads.set()
                break

        time.sleep(poll_s)


def get_current_time() -> str:
    return time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="UltraArm P1 aging test (joint limits loop).")
    parser.add_argument("--port", default="COM9", help="Serial port, e.g. COM11")
    parser.add_argument("--baud", type=int, default=1_000_000, help="Baudrate")
    parser.add_argument("--speed", type=int, default=80, help="set_angles speed")
    parser.add_argument("--poll", type=float, default=0.2, help="Monitor poll interval seconds")
    parser.add_argument("--cycles", type=int, default=0, help="Cycle count; 0 means infinite")
    parser.add_argument("--report-dir", default="test_report", help="Report directory")
    parser.add_argument(
        "--delta-xyz",
        type=float,
        default=20.0,
        help="笛卡尔 X/Y/Z 单轴对称运动步长（mm），会自动夹在软限位内",
    )
    parser.add_argument(
        "--delta-rx",
        type=float,
        default=20.0,
        help="笛卡尔 Rx 对称运动步长（°）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    os.makedirs(args.report_dir, exist_ok=True)

    file_name = f"{get_current_time()}_P1_aging_movement.xlsx"
    save_path = os.path.join(os.getcwd(), args.report_dir, file_name)

    wb = Workbook()
    ws_movement = wb.active
    ws_movement.title = "关节运动"
    ws_movement.append(
        [
            "轮次",
            "关节",
            "负向用时(s)",
            "正向用时(s)",
            "负向失败",
            "正向失败",
            "负向目标角度",
            "正向目标角度",
        ]
    )

    ws_coords = wb.create_sheet("坐标运动")
    ws_coords.append(
        [
            "轮次",
            "轴",
            "负向用时(s)",
            "正向用时(s)",
            "负向失败",
            "正向失败",
            "负向目标坐标",
            "正向目标坐标",
            "负向反馈坐标",
            "正向反馈坐标",
            "本轮基准坐标",
        ]
    )

    ws_monitor_summary = wb.create_sheet("监控汇总")
    ws_monitor_summary.append(["项目", "数值"])

    ws_monitor_events = wb.create_sheet("监控事件")
    ws_monitor_events.append(
        ["时间", "错误信息", "运行状态", "电机使能", "异常计数"]
    )

    movement_stats = {
        "angles_attempts": 0,
        "angles_failed": 0,
        "coords_attempts": 0,
        "coords_failed": 0,
    }
    monitor_stats = {
        "poll_count": 0,
        "error_info_invalid": 0,
        "run_status_invalid": 0,
        "motor_enable_invalid": 0,
        "monitor_exceptions": 0,
    }
    monitor_events: list[list[Any]] = []

    mc = UltraArmP1(args.port, args.baud, debug=1)
    _mc_holder["mc"] = mc
    _install_signal_handlers()
    logger.info(
        "停止测试：按 Ctrl+C；若无效可试 Ctrl+Break（Windows），或关闭此终端窗口结束进程。"
    )
    logger.info(
        "P1 老化测试启动：串口=%s，波特率=%s，速度=%s，循环=%s，"
        "关节到位±%s°，坐标 XYZ±%s mm、Rx±%s°，坐标步长 XYZ=%s mm、Rx=%s°",
        args.port,
        args.baud,
        args.speed,
        args.cycles,
        ANGLE_TOLERANCE_DEG,
        COORD_TOLERANCE_MM,
        COORD_TOLERANCE_RX_DEG,
        args.delta_xyz,
        args.delta_rx,
    )

    try:
        mc.set_angles(INIT_ANGLES, args.speed)
        init_ok = wait_motion_done(mc, timeout_sec=20.0)
        if not init_ok:
            logger.info("初始化阶段已停止或超时，不启动运动/监控线程")

        t_move: threading.Thread | None = None
        t_monitor: threading.Thread | None = None
        if init_ok:
            t_move = threading.Thread(
                target=movement_thread,
                name="P1MoveThread",
                args=(
                    mc,
                    args.speed,
                    ws_movement,
                    ws_coords,
                    wb,
                    save_path,
                    movement_stats,
                    args.cycles,
                    args.delta_xyz,
                    args.delta_rx,
                ),
            )
            t_monitor = threading.Thread(
                target=monitor_thread,
                name="P1MonitorThread",
                args=(mc, args.poll, monitor_stats, monitor_events),
            )

            t_move.start()
            t_monitor.start()

            try:
                _join_threads_until_done(t_move, t_monitor)
            except KeyboardInterrupt:
                logger.info("收到 KeyboardInterrupt，正在停止测试…")
                _request_shutdown()
                _join_threads_until_done(t_move, t_monitor, max_wait_sec=45.0)

    except KeyboardInterrupt:
        logger.info("收到 KeyboardInterrupt，正在停止测试…")
        _request_shutdown()
    except Exception as e:
        logger.error("测试运行异常：%s", e)
        stop_threads.set()
    finally:
        stop_threads.set()

        # Write monitor summary/events once threads finish.
        ws_monitor_summary.append(["串口", args.port])
        ws_monitor_summary.append(["波特率", args.baud])
        ws_monitor_summary.append(["速度", args.speed])
        ws_monitor_summary.append(["监控周期(s)", args.poll])
        ws_monitor_summary.append(["循环次数(0=无限)", args.cycles])
        ws_monitor_summary.append(["起始关节角", str(INIT_ANGLES)])
        ws_monitor_summary.append(["笛卡尔前置关节角", str(COORDS_INIT_ANGLES)])
        ws_monitor_summary.append(["关节到位容差(°)", ANGLE_TOLERANCE_DEG])
        ws_monitor_summary.append(["坐标 XYZ 容差(mm)", COORD_TOLERANCE_MM])
        ws_monitor_summary.append(["坐标 Rx 容差(°)", COORD_TOLERANCE_RX_DEG])
        ws_monitor_summary.append(["坐标步长 XYZ(mm)", args.delta_xyz])
        ws_monitor_summary.append(["坐标步长 Rx(°)", args.delta_rx])
        ws_monitor_summary.append(["关节发送次数", movement_stats["angles_attempts"]])
        ws_monitor_summary.append(["关节失败次数", movement_stats["angles_failed"]])
        ws_monitor_summary.append(["坐标发送次数", movement_stats["coords_attempts"]])
        ws_monitor_summary.append(["坐标失败次数", movement_stats["coords_failed"]])
        ws_monitor_summary.append(["监控轮询次数", monitor_stats["poll_count"]])
        ws_monitor_summary.append(["错误信息无效次数", monitor_stats["error_info_invalid"]])
        ws_monitor_summary.append(["运行状态无效次数", monitor_stats["run_status_invalid"]])
        ws_monitor_summary.append(["电机使能无效次数", monitor_stats["motor_enable_invalid"]])
        ws_monitor_summary.append(["监控异常次数", monitor_stats["monitor_exceptions"]])

        for row in monitor_events:
            ws_monitor_events.append(row)

        try:
            wb.save(save_path)
            logger.info("报告已保存：%s", save_path)
        except Exception as e:
            logger.error("保存报告失败：%s", e)

        try:
            mc.close()
        except Exception:
            pass
        _mc_holder["mc"] = None


if __name__ == "__main__":
    main()
