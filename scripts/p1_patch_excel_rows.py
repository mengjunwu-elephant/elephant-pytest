# -*- coding: utf-8 -*-
"""补全/修正 UltraArm P1 Excel 用例行（不重复创建已有 sheet）。"""
from __future__ import annotations

import os

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAIN = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")
ATT = os.path.join(BASE, "test_data", "UltraArm_P1_Attachments.xlsx")


def _ensure_rows(wb, sheet: str, headers: list[str], rows: list[list]) -> None:
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(headers)
    else:
        ws = wb[sheet]
    existing = ws.max_row - 1
    if existing >= len(rows):
        return
    for row in rows[existing:]:
        ws.append(row)


def main() -> None:
    wb = load_workbook(MAIN)
    _ensure_rows(
        wb,
        "get_run_status",
        ["ID", "title", "api", "parameter", "expect_data", "test_type"],
        [
            [1, "静止时读取运行状态", "get_run_status", None, 0, "normal"],
            [2, "运动中读取运行状态", "get_run_status", None, 1, "normal1"],
        ],
    )
    _ensure_rows(
        wb,
        "get_base_io_state",
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [
            [1, "读取底座IO引脚1", "get_base_io_state", 1, 0, "normal"],
            [2, "引脚号越界", "get_base_io_state", 0, None, "exception"],
        ],
    )
    _ensure_rows(
        wb,
        "get_end_io_state",
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [
            [1, "读取末端IO引脚3", "get_end_io_state", 3, 0, "normal"],
            [2, "引脚号越界", "get_end_io_state", 1, None, "exception"],
        ],
    )
    wb.save(MAIN)

    wb_att = load_workbook(ATT)
    _ensure_rows(
        wb_att,
        "set_gripper_parameter",
        ["ID", "title", "api", "parameter", "value", "expect_data", "test_type"],
        [
            [1, "设置夹爪参数", "set_gripper_parameter", 1, 100, 1, "normal"],
            [2, "地址越界", "set_gripper_parameter", 0, 100, None, "exception"],
        ],
    )
    wb_att.save(ATT)
    print("Excel patches applied.")


if __name__ == "__main__":
    main()
