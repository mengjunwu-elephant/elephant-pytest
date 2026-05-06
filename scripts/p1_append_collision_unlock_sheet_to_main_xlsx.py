# -*- coding: utf-8 -*-
"""将 test_data/UltraArm_P1_collision_unlock.xlsx 中的 collision_unlock 表追加到
test_data/UltraArm_P1.xlsx（主表无同名 sheet 时）。

在关闭 Excel、无其他进程占用主表后执行：
  python scripts/p1_append_collision_unlock_sheet_to_main_xlsx.py
成功后可只保留主表，并把 settings 中 UltraArmP1Base.COLLISION_UNLOCK_DATA_FILE
改为主表路径以统一数据源（可选）。"""
from __future__ import annotations

import os
import shutil
import sys

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAIN = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")
AUX = os.path.join(BASE, "test_data", "UltraArm_P1_collision_unlock.xlsx")
SHEET = "collision_unlock"


def main() -> int:
    if not os.path.isfile(MAIN) or not os.path.isfile(AUX):
        print("缺少主表或副表文件", file=sys.stderr)
        return 1
    wb_main = load_workbook(MAIN)
    if SHEET in wb_main.sheetnames:
        print(f"主表已存在 {SHEET!r}，未修改")
        return 0
    wb_aux = load_workbook(AUX, read_only=True, data_only=True)
    if SHEET not in wb_aux.sheetnames:
        print("副表缺少 collision_unlock", file=sys.stderr)
        return 1
    sh = wb_aux[SHEET]
    nws = wb_main.create_sheet(SHEET)
    for row in sh.iter_rows():
        nws.append([c.value for c in row])
    wb_aux.close()
    bak = MAIN + ".bak"
    shutil.copy2(MAIN, bak)
    wb_main.save(MAIN)
    print(f"已追加 {SHEET} 到主表，备份: {bak}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
