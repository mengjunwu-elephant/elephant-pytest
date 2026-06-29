# -*- coding: utf-8
"""按 UltraArmP1Base 限位更新运动接口 Excel 边界用例。"""
from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")

MIN_ANGLES = [-165, -18, 89, -179]
MAX_ANGLES = [165, 85, 200, 179]
MIN_COORDS = [-350, -362.43, -186.265, -180.0]
MAX_COORDS = [360.43, 362.43, 93.44, 180.0]

# 越界：角度 ±1°，坐标略超出限位
ANG_OOR_MIN = [-166, -19, 88, -180]
ANG_OOR_MAX = [166, 86, 201, 180]
COORD_OOR = [-351, -363, -187, -181]
COORD_OOR_MAX = [361, 363, 94, 181]


def _set_by_id(ws, id_col: int, target_id: int, updates: dict[int, object]) -> None:
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, id_col).value == target_id:
            for col, val in updates.items():
                ws.cell(r, col, val)
            return


def _col_map(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def update_set_angle(ws) -> None:
    cm = _col_map(ws)
    for jid in range(1, 5):
        _set_by_id(ws, cm["ID"], 2 * jid - 1, {cm["angle"]: MIN_ANGLES[jid - 1]})
        _set_by_id(ws, cm["ID"], 2 * jid, {cm["angle"]: MAX_ANGLES[jid - 1]})
        _set_by_id(ws, cm["ID"], 24 + 2 * jid - 1, {cm["angle"]: ANG_OOR_MIN[jid - 1]})
        _set_by_id(ws, cm["ID"], 24 + 2 * jid, {cm["angle"]: ANG_OOR_MAX[jid - 1]})


def update_set_angles(ws) -> None:
    cm = _col_map(ws)
    templates = [
        (1, f"[{MIN_ANGLES[0]},0,90,0]"),
        (2, f"[{MAX_ANGLES[0]},0,90,0]"),
        (3, f"[0,{MIN_ANGLES[1]},110,0]"),
        (4, f"[0,{MAX_ANGLES[1]},90,0]"),
        (5, f"[0,0,{MIN_ANGLES[2]},0]"),
        (6, f"[0,50,{MAX_ANGLES[2]},0]"),
        (7, f"[0,0,90,{MIN_ANGLES[3]}]"),
        (8, f"[0,0,90,{MAX_ANGLES[3]}]"),
    ]
    for row_id, angles in templates:
        _set_by_id(ws, cm["ID"], row_id, {cm["angles"]: angles})
    exc = [
        (25, f"[{ANG_OOR_MIN[0]},0,90,0]"),
        (26, f"[{ANG_OOR_MAX[0]},0,90,0]"),
        (27, f"[0,{ANG_OOR_MIN[1]},90,0]"),
        (28, f"[0,{ANG_OOR_MAX[1]},90,0]"),
        (29, f"[0,0,{ANG_OOR_MIN[2]},0]"),
        (30, f"[0,0,{ANG_OOR_MAX[2]},0]"),
        (31, f"[0,0,90,{ANG_OOR_MIN[3]}]"),
        (32, f"[0,0,90,{ANG_OOR_MAX[3]}]"),
    ]
    for row_id, angles in exc:
        _set_by_id(ws, cm["ID"], row_id, {cm["angles"]: angles})


def update_set_coords(ws) -> None:
    cm = _col_map(ws)
    normals = [
        (1, f"[{MAX_COORDS[0]}, 0.0, -17, 0.0]"),
        (2, f"[{MIN_COORDS[0]}, 0.0, -17, 0.0]"),
        (3, f"[260, {MAX_COORDS[1]}, -17, 0.0]"),
        (4, f"[260, {MIN_COORDS[1]}, -17, 0.0]"),
        (5, f"[260, 0.0, {MAX_COORDS[2]}, 0.0]"),
        (6, f"[260, 0.0, {MIN_COORDS[2]}, 0.0]"),
        (7, f"[260, 0.0, -17, {MAX_COORDS[3]}]"),
        (8, f"[260, 0.0, -17, {MIN_COORDS[3]}]"),
    ]
    for row_id, coords in normals:
        _set_by_id(ws, cm["ID"], row_id, {cm["coords"]: coords})
    exc = [
        (25, f"[{COORD_OOR[0]},0.00,-17,-30]"),
        (26, f"[{COORD_OOR_MAX[0]},0.00,-17,-30]"),
        (27, f"[83.50,{COORD_OOR_MAX[1]},-17,-30]"),
        (28, f"[83.50,{COORD_OOR[1]},-17,-30]"),
        (29, f"[83.50,0.00,{COORD_OOR[2]},-30]"),
        (30, f"[83.50,0.00,{COORD_OOR_MAX[2]},-30]"),
    ]
    for row_id, coords in exc:
        _set_by_id(ws, cm["ID"], row_id, {cm["coords"]: coords})


def update_set_coord(ws) -> None:
    cm = _col_map(ws)
    axis_vals = [
        (1, "X", MAX_COORDS[0]),
        (2, "X", MIN_COORDS[0]),
        (3, "Y", MAX_COORDS[1]),
        (4, "Y", MIN_COORDS[1]),
        (5, "Z", MAX_COORDS[2]),
        (6, "Z", MIN_COORDS[2]),
        (7, "R", MAX_COORDS[3]),
        (8, "R", MIN_COORDS[3]),
    ]
    for row_id, axis, val in axis_vals:
        _set_by_id(ws, cm["ID"], row_id, {cm["axis"]: axis, cm["coord"]: val})
    exc_axis = [
        (25, 1, COORD_OOR[0]),
        (26, 1, COORD_OOR_MAX[0]),
        (27, 2, COORD_OOR_MAX[1]),
        (28, 2, COORD_OOR[1]),
        (29, 3, COORD_OOR_MAX[2]),
        (30, 3, COORD_OOR[2]),
    ]
    for row_id, axis, val in exc_axis:
        _set_by_id(ws, cm["ID"], row_id, {cm["axis"]: axis, cm["coord"]: val})


def main() -> None:
    shutil.copy2(XLSX, XLSX + ".bak_limits")
    wb = load_workbook(XLSX)
    update_set_angle(wb["set_angle"])
    update_set_angles(wb["set_angles"])
    update_set_coords(wb["set_coords"])
    update_set_coord(wb["set_coord"])
    wb.save(XLSX)
    print("Updated set_angle / set_angles / set_coords / set_coord limit rows.")


if __name__ == "__main__":
    main()
