# -*- coding: utf-8 -*-
"""为 UltraArm P1 补全/追加 Excel 测试 sheet（主表 + 附件表）。

用法（关闭 Excel 后）:
  D:\\python\\python.exe scripts/p1_bootstrap_test_sheets.py
"""
from __future__ import annotations

import os
import shutil
import sys

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAIN = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")
ATT = os.path.join(BASE, "test_data", "UltraArm_P1_Attachments.xlsx")
AUX = os.path.join(BASE, "test_data", "UltraArm_P1_collision_unlock.xlsx")

# sheet_name -> (headers, rows)
MAIN_SHEETS: dict[str, tuple[list[str], list[list]]] = {
    "get_run_status": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [
            [1, "静止时读取运行状态", "get_run_status", 0, "normal"],
            [2, "运动中读取运行状态", "get_run_status", 1, "moving"],
        ],
    ),
    "get_zero_calibration_state": (
        ["ID", "title", "api", "joint", "expect_data", "test_type"],
        [[1, "读取零位校准状态", "get_zero_calibration_state", 0, "[1,1,1,1]", "normal"]],
    ),
    "get_base_io_input": (
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [[1, "读取底座IO引脚1", "get_base_io_state", 1, 0, "normal"]],
    ),
    "set_coords_max_speed": (
        ["ID", "title", "api", "coords", "expect_data", "test_type"],
        [
            [1, "最大速度坐标运动", "set_coords_max_speed", "[170, 0.0, 41.1, 0.0]", "ok", "normal"],
        ],
    ),
    "clear_error_status": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "无错误时清除", "clear_error_status", "OK", "normal2"]],
    ),
    "coord_inverse_solution": (
        ["ID", "title", "api", "coords", "expect_data", "test_type", "tol"],
        [
            [1, "坐标逆解", "coord_inverse_solution", "[170, 0.0, 41.1, 0.0]", "[0,20,110,0]", "normal", 1.0],
        ],
    ),
    "angle_correct_solution": (
        ["ID", "title", "api", "angles", "expect_data", "test_type", "tol"],
        [
            [1, "角度正解", "angle_correct_solution", "[0, 20, 110, 0]", "[170,0,41,0]", "normal", 2.0],
        ],
    ),
    "forced_reset_zero": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "强制回零", "forced_reset_zero", "ok", "normal"]],
    ),
    "set_preview_mode": (
        ["ID", "title", "api", "coords", "expect_data", "test_type"],
        [[1, "轨迹预览", "set_preview_mode", "[170, 0.0, 41.1, 0.0]", "ok", "normal"]],
    ),
    "set_pwm_laser_mode": (
        ["ID", "title", "api", "state", "expect_data", "test_type"],
        [
            [1, "打开激光模式", "set_pwm_laser_mode", 1, "ok", "normal"],
            [2, "关闭激光模式", "set_pwm_laser_mode", 0, "ok", "normal_off"],
        ],
    ),
    "set_pwm_laser": (
        ["ID", "title", "api", "p_value", "expect_data", "test_type"],
        [[1, "设置激光PWM档位", "set_pwm_laser", 128, "ok", "normal"]],
    ),
    "set_pwm_custom_mode": (
        ["ID", "title", "api", "state", "expect_data", "test_type"],
        [
            [1, "打开自定义PWM", "set_pwm_custom_mode", 1, "ok", "normal"],
            [2, "关闭自定义PWM", "set_pwm_custom_mode", 0, "ok", "normal_off"],
        ],
    ),
    "set_pwm_custom": (
        ["ID", "title", "api", "p_value", "expect_data", "test_type"],
        [[1, "设置自定义PWM档位", "set_pwm_custom", 100, "ok", "normal"]],
    ),
    "get_pwm_status": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取PWM状态", "get_pwm_status", "[0,0,0,0]", "normal"]],
    ),
    "set_color": (
        ["ID", "title", "api", "r", "g", "b", "expect_data", "test_type"],
        [[1, "设置RGB红色", "set_color", 255, 0, 0, "ok", "normal"]],
    ),
    "set_conveyor_control": (
        ["ID", "title", "api", "state", "direction", "speed", "distance", "expect_data", "test_type"],
        [[1, "传送带前进短距", "set_conveyor_control", 1, 0, 100, 100, "ok", "normal"]],
    ),
    "set_conveyor_stop": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "传送带停止", "set_conveyor_stop", "ok", "normal"]],
    ),
    "get_all_base_io_states": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取全部底座IO", "get_all_base_io_states", None, "normal"]],
    ),
    "get_base_io_state": (
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [[1, "读取底座IO引脚1", "get_base_io_state", 1, 0, "normal"]],
    ),
    "get_all_end_io_states": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取全部末端IO", "get_all_end_io_states", None, "normal"]],
    ),
    "get_end_io_state": (
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [[1, "读取末端IO引脚3", "get_end_io_state", 3, 0, "normal"]],
    ),
    "set_end_button_enable": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "末端按钮使能", "set_end_button_enable", "ok", "normal"]],
    ),
    "set_end_button_disable": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "末端按钮失能", "set_end_button_disable", "ok", "normal"]],
    ),
    "get_end_button_state": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取末端按钮状态", "get_end_button_state", 0, "normal"]],
    ),
    "set_collision_threshold": (
        ["ID", "title", "api", "joint_id", "threshold", "expect_data", "test_type", "restore_threshold"],
        [
            [1, "设置J1碰撞阈值", "set_collision_threshold", 1, 1.0, "ok", "normal", 0.5],
            [2, "关节号越界", "set_collision_threshold", 5, 1.0, None, "exception", None],
        ],
    ),
    "get_collision_threshold": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取碰撞阈值", "get_collision_threshold", "[0.5,0.5,0.5,0.5]", "normal"]],
    ),
    "set_robot_id": (
        ["ID", "title", "api", "robot_id", "expect_data", "test_type", "restore_robot_id"],
        [[1, "设置机器码", "set_robot_id", "001", "ok", "normal", None]],
    ),
    "get_robot_id": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取机器码", "get_robot_id", "001", "normal"]],
    ),
    "get_communication_mode": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取通信模式", "get_communication_mode", "Uart0", "normal"]],
    ),
    "set_uart1_communication": (
        ["ID", "title", "api", "state", "expect_data", "test_type", "restore_state"],
        [
            [1, "打开串口1通信", "set_uart1_communication", 1, "ok", "normal", 0],
            [2, "状态越界", "set_uart1_communication", 2, None, "exception", None],
        ],
    ),
    "get_wifi_ip": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取WiFi IP", "get_wifi_ip", None, "normal"]],
    ),
    "get_bluetooth_mac": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取蓝牙MAC", "get_bluetooth_mac", None, "normal"]],
    ),
    "get_wifi_signal_strength": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取WiFi信号", "get_wifi_signal_strength", None, "normal"]],
    ),
    "get_bluetooth_signal_strength": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取蓝牙信号", "get_bluetooth_signal_strength", None, "normal"]],
    ),
    "get_system_screen_version": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取屏幕主版本", "get_system_screen_version", 1, "normal"]],
    ),
    "get_modify_screen_version": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取屏幕更正版本", "get_modify_screen_version", 0, "normal"]],
    ),
    "check_sd_card": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "检查SD卡", "check_sd_card", "ok", "normal"]],
    ),
    "get_sd_card_space": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "读取SD卡空间", "get_sd_card_space", None, "normal"]],
    ),
    "play_gcode_file": (
        ["ID", "title", "api", "filename", "expect_data", "test_type"],
        [[1, "播放轨迹文件", "play_gcode_file", "test.gcode", None, "normal"]],
    ),
    "set_communication_baud_rate": (
        ["ID", "title", "api", "baud_rate", "expect_data", "test_type"],
        [[1, "波特率越界", "set_communication_baud_rate", 9600, None, "exception"]],
    ),
    "set_wifi_password": (
        ["ID", "title", "api", "wifi_name", "password", "expect_data", "test_type"],
        [[1, "密码长度不足", "set_wifi_password", "test_wifi", "123", None, "exception"]],
    ),
    "download_firmware_sd": (
        ["ID", "title", "api", "filename", "expect_data", "test_type"],
        [[1, "固件下载占位", "download_firmware_sd", "fw.bin", None, "skip"]],
    ),
    "upgrade_restart": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "固件升级重启占位", "upgrade_restart", "ok", "skip"]],
    ),
    "finish_firmware_upgrade": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "结束固件下载占位", "finish_firmware_upgrade", "ok", "skip"]],
    ),
    "receive_485_data": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "接收485数据", "receive_485_data", None, "normal"]],
    ),
}

ATT_SHEETS: dict[str, tuple[list[str], list[list]]] = {
    "close_laser": (
        ["ID", "title", "api", "expect_data", "test_type"],
        [[1, "关闭激光", "set_pwm_laser_mode", 0, "ok", "normal"]],
    ),
    "set_pump_state": (
        ["ID", "title", "api", "pump_state", "expect_data", "test_type"],
        [
            [1, "打开吸泵", "set_pump_state", 0, "ok", "normal"],
            [2, "释放吸泵", "set_pump_state", 1, "ok", "normal_release"],
            [3, "关闭吸泵", "set_pump_state", 2, "ok", "normal_close"],
            [4, "状态越界", "set_pump_state", 3, None, "exception"],
        ],
    ),
}


def _append_sheet(wb, name: str, headers: list[str], rows: list[list]) -> bool:
    if name in wb.sheetnames:
        return False
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return True


def _merge_collision(wb) -> bool:
    if "collision_unlock" in wb.sheetnames:
        return False
    if not os.path.isfile(AUX):
        return False
    wb_aux = load_workbook(AUX, read_only=True, data_only=True)
    if "collision_unlock" not in wb_aux.sheetnames:
        wb_aux.close()
        return False
    sh = wb_aux["collision_unlock"]
    nws = wb.create_sheet("collision_unlock")
    for row in sh.iter_rows():
        nws.append([c.value for c in row])
    wb_aux.close()
    return True


def _bootstrap(path: str, sheets: dict[str, tuple[list[str], list[list]]]) -> list[str]:
    wb = load_workbook(path)
    added: list[str] = []
    for name, (headers, rows) in sheets.items():
        if _append_sheet(wb, name, headers, rows):
            added.append(name)
    if path == MAIN:
        if _merge_collision(wb):
            added.append("collision_unlock")
    bak = path + ".bak_bootstrap"
    if os.path.isfile(path):
        shutil.copy2(path, bak)
    wb.save(path)
    return added


def main() -> int:
    for p in (MAIN, ATT):
        if not os.path.isfile(p):
            print(f"缺少文件: {p}", file=sys.stderr)
            return 1
    main_added = _bootstrap(MAIN, MAIN_SHEETS)
    att_added = _bootstrap(ATT, ATT_SHEETS)
    print(f"主表新增 sheet: {main_added}")
    print(f"附件表新增 sheet: {att_added}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
