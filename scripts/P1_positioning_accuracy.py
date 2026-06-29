# -*- coding: utf-8 -*-
"""
UltraArm P1：绝对定位精度 & 重复定位精度 demo（结果写入 Excel）。

使用接口：set_angles / set_angles + get_angles_info；set_coords / set_coords + get_coords_info。
运动结束后根据控制器反馈计算误差（非外部测量设备）。

重复定位精度采用「远离点 → 测定点」往返：每次先离开测定点，再重新逼近并采样。
绝对定位精度同样采用该路径，在测定点统计指令与反馈的偏差（均值、方差、RMSE 等）。

默认每个目标点往返 100 次（``--repeat 100``），绝对/重复定位均输出明细与汇总表。

运行示例（在项目根目录）::

    python scripts/P1_positioning_accuracy.py --port COM10 --baud 1000000

默认会在 scripts 目录下生成 ``P1_positioning_accuracy_<时间戳>.xlsx``。
"""

from __future__ import annotations

import argparse
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from pymycobot import UltraArmP1

# ---------------------------------------------------------------------------
# 关节 / 笛卡尔限位（按你的规格；请在不确定处与文档/实机核对）
# ---------------------------------------------------------------------------
JOINT_LIMITS_DEG = (
    (-165.0, 165.0),  # J1
    (-18.0, 85.0),  # J2
    (89.0, 200.0),  # J3（与 pymycobot set_angles 校验一致：89 ~ 190）
    (-179.0, 179.0),  # J4
)

# 笛卡尔运动前建议姿态（与 settings.UltraArmP1Base.coords_init_angles、接口测试一致）。
# 若在关节测试后直接 set_coords，易从「奇异/耦合」关节域出发，固件可能返回 LimitError:6。
# pymycobot ultraArmP1._response：LimitError[6] => "J2/J3 joint coupling limit."
COORDS_INIT_ANGLES: list[float] = [0.0, 10.0, 110.0, 0.0]

# [x, y, z, rx] —— P1 示教坐标与仓库用例一致为 4 维
COORD_LIMITS = (
    (-350.7, 362.43),  # X
    (-362.43, 362.43),  # Y
    (-186.265, 93.44),  # Z
    (-180.0, 180.0),  # Rx
)

# 默认测试指令位姿（关节空间）：需在各自限位内且尽量为实机可达姿态，可按现场修改
DEFAULT_ANGLE_TARGETS: list[list[float]] = [
    [0.0, 30.0, 120.0, 0.0],
    [-45.0, 40.0, 150.0, 15.0],
    [60.0, 10.0, 100.0, -30.0],
]

# 默认笛卡尔测试点（需在坐标限位内且可达；若不可达请换成你们标定过的点）
DEFAULT_COORD_TARGETS: list[list[float]] = [
    [250, 20, 30, 10],
    [270.0, 50.0, 40.0, 30.0],
    [230.0, -60.0, 50.0, -45.0],
]

# 重复定位精度：每次先运动到「远离点」，再运动到「测定点」并采样（ departure → approach ）
DEFAULT_ANGLE_AWAY: list[float] = list(COORDS_INIT_ANGLES)
DEFAULT_COORD_AWAY: list[float] = [230.0, 0.0, -30.0, 0.0]

# 运动结束后读反馈：超时/空包时重试
READ_MAX_ATTEMPTS = 10
READ_RETRY_INTERVAL_SEC = 0.2
READ_SETTLE_SEC = 0.15


def _in_joint_limits(angles: Sequence[float]) -> bool:
    for v, (lo, hi) in zip(angles, JOINT_LIMITS_DEG):
        if not (lo <= v <= hi):
            return False
    return True
 

def _in_coord_limits(coords: Sequence[float]) -> bool:
    for v, (lo, hi) in zip(coords, COORD_LIMITS):
        if not (lo <= v <= hi):
            return False
    return True


def move_to_coords_init(mc: UltraArmP1, speed: int) -> None:
    """先回到便于笛卡尔规划的姿态，减轻 J2/J3 耦合触发 LimitError:6。"""
    mc.set_angles(list(COORDS_INIT_ANGLES), speed)
    wait_motion_done(mc)


def wait_motion_done(mc: UltraArmP1, timeout_sec: float = 120.0) -> None:
    deadline = time.monotonic() + float(timeout_sec)
    time.sleep(0.3)
    while True:
        if time.monotonic() > deadline:
            raise TimeoutError(f"等待运动结束超时（{timeout_sec}s）")
        try:
            status = mc.get_run_status()
        except Exception:
            time.sleep(0.1)
            continue
        if status == -1:
            time.sleep(0.1)
            continue
        if not status:
            break
        time.sleep(0.1)
    time.sleep(0.3)


def _float_row(xs: Sequence[Any]) -> list[float]:
    return [float(x) for x in xs]


def _is_pose_row(data: Any, *, expected_len: int = 4) -> bool:
    """pymycobot 读失败时常返回 -1 或不可迭代标量，需过滤后再解析。"""
    if data is None or data == -1:
        return False
    if isinstance(data, (int, float)) and not isinstance(data, bool):
        return False
    try:
        row = list(data)
    except TypeError:
        return False
    if len(row) != expected_len:
        return False
    try:
        _float_row(row)
    except (TypeError, ValueError):
        return False
    return True


def read_angles_info(
    mc: UltraArmP1,
    *,
    max_attempts: int | None = None,
    retry_interval_sec: float | None = None,
    settle_sec: float | None = None,
) -> list[float]:
    """读取关节角；无效/超时则间隔重试，避免把 -1 当序列解析。"""
    attempts = READ_MAX_ATTEMPTS if max_attempts is None else max_attempts
    interval = READ_RETRY_INTERVAL_SEC if retry_interval_sec is None else retry_interval_sec
    settle = READ_SETTLE_SEC if settle_sec is None else settle_sec
    if settle > 0:
        time.sleep(settle)
    last: Any = None
    for attempt in range(1, attempts + 1):
        try:
            last = mc.get_angles_info()
        except Exception as exc:
            last = exc
        if _is_pose_row(last):
            return _float_row(last)
        if attempt < attempts:
            print(
                f"[WARN] get_angles_info 无效({last!r})，"
                f"{attempt}/{attempts} 次，{interval}s 后重试…",
                file=sys.stderr,
            )
            time.sleep(interval)
    raise TimeoutError(
        f"get_angles_info 读取失败（已重试 {attempts} 次），最后一次: {last!r}"
    )


def read_coords_info(
    mc: UltraArmP1,
    *,
    max_attempts: int | None = None,
    retry_interval_sec: float | None = None,
    settle_sec: float | None = None,
) -> list[float]:
    """读取笛卡尔坐标；无效/超时则间隔重试。"""
    attempts = READ_MAX_ATTEMPTS if max_attempts is None else max_attempts
    interval = READ_RETRY_INTERVAL_SEC if retry_interval_sec is None else retry_interval_sec
    settle = READ_SETTLE_SEC if settle_sec is None else settle_sec
    if settle > 0:
        time.sleep(settle)
    last: Any = None
    for attempt in range(1, attempts + 1):
        try:
            last = mc.get_coords_info()
        except Exception as exc:
            last = exc
        if _is_pose_row(last):
            return _float_row(last)
        if attempt < attempts:
            print(
                f"[WARN] get_coords_info 无效({last!r})，"
                f"{attempt}/{attempts} 次，{interval}s 后重试…",
                file=sys.stderr,
            )
            time.sleep(interval)
    raise TimeoutError(
        f"get_coords_info 读取失败（已重试 {attempts} 次），最后一次: {last!r}"
    )


def _poses_differ(a: Sequence[float], b: Sequence[float], *, tol: float = 1e-6) -> bool:
    return any(abs(float(x) - float(y)) > tol for x, y in zip(a, b))


def _repeat_position_stats(values: Sequence[float]) -> dict[str, float]:
    """重复定位：反馈值的均值、方差、标准差、3σ、极差等。"""
    col = list(values)
    n = len(col)
    if n == 0:
        return {
            "均值": 0.0,
            "方差": 0.0,
            "标准差": 0.0,
            "3σ": 0.0,
            "最小值": 0.0,
            "最大值": 0.0,
            "极差": 0.0,
            "相对均值最大偏差": 0.0,
        }
    mu = statistics.fmean(col)
    var = statistics.pvariance(col) if n > 1 else 0.0
    sigma = statistics.pstdev(col) if n > 1 else 0.0
    lo, hi = min(col), max(col)
    return {
        "均值": mu,
        "方差": var,
        "标准差": sigma,
        "3σ": 3.0 * sigma,
        "最小值": lo,
        "最大值": hi,
        "极差": hi - lo,
        "相对均值最大偏差": max(abs(x - mu) for x in col),
    }


def _absolute_error_stats(errors: Sequence[float]) -> dict[str, float]:
    """绝对定位：偏差均值、偏差方差、均方根误差、最大绝对误差等。"""
    err = list(errors)
    n = len(err)
    if n == 0:
        return {
            "偏差均值": 0.0,
            "偏差方差": 0.0,
            "偏差标准差": 0.0,
            "均方根误差": 0.0,
            "最大绝对误差": 0.0,
        }
    mu = statistics.fmean(err)
    var = statistics.pvariance(err) if n > 1 else 0.0
    rmse = statistics.fmean(e * e for e in err) ** 0.5
    return {
        "偏差均值": mu,
        "偏差方差": var,
        "偏差标准差": statistics.pstdev(err) if n > 1 else 0.0,
        "均方根误差": rmse,
        "最大绝对误差": max(abs(e) for e in err),
    }


def run_absolute_angles(
    mc: UltraArmP1,
    speed: int,
    targets: list[list[float]],
    repeat_n: int,
    away: Sequence[float] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """绝对定位：远离点 → 测定点，循环 repeat_n 次后在测定点采样并统计指令偏差。"""
    detail: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    axis_names = ("J1", "J2", "J3", "J4")
    awayf = _float_row(away if away is not None else DEFAULT_ANGLE_AWAY)
    if not _in_joint_limits(awayf):
        raise ValueError(f"关节远离点超出限位: {awayf}")

    for pi, cmd in enumerate(targets):
        cmdf = _float_row(cmd)
        if not _in_joint_limits(cmdf):
            raise ValueError(f"关节目标 {pi} 超出限位: {cmdf}")
        if not _poses_differ(awayf, cmdf):
            raise ValueError(f"关节测定点 {pi + 1} 与远离点相同，无法评估绝对定位: {cmdf}")
        errors_by_axis: list[list[float]] = [[], [], [], []]
        for k in range(repeat_n):
            mc.set_angles(awayf, speed)
            wait_motion_done(mc)
            mc.set_angles(cmdf, speed)
            wait_motion_done(mc)
            act = read_angles_info(mc)
            err = [a - c for a, c in zip(act, cmdf)]
            for j in range(4):
                errors_by_axis[j].append(err[j])
            detail.append(
                {
                    "目标点序号": pi + 1,
                    "重复序号": k + 1,
                    "运动路径": "远离点→测定点",
                    "J1远离点(°)": awayf[0],
                    "J2远离点(°)": awayf[1],
                    "J3远离点(°)": awayf[2],
                    "J4远离点(°)": awayf[3],
                    "J1测定点指令(°)": cmdf[0],
                    "J2测定点指令(°)": cmdf[1],
                    "J3测定点指令(°)": cmdf[2],
                    "J4测定点指令(°)": cmdf[3],
                    "J1反馈(°)": act[0],
                    "J2反馈(°)": act[1],
                    "J3反馈(°)": act[2],
                    "J4反馈(°)": act[3],
                    "J1误差(°)": err[0],
                    "J2误差(°)": err[1],
                    "J3误差(°)": err[2],
                    "J4误差(°)": err[3],
                    "四轴最大绝对误差(°)": max(abs(x) for x in err),
                }
            )

        for j, name in enumerate(axis_names):
            row = {
                "目标点序号": pi + 1,
                "关节": name,
                "测定点指令(°)": cmdf[j],
            }
            row.update(_absolute_error_stats(errors_by_axis[j]))
            summary.append(row)

    return detail, summary


def run_repeat_angles(
    mc: UltraArmP1,
    speed: int,
    targets: list[list[float]],
    repeat_n: int,
    away: Sequence[float] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """重复定位：远离点 → 测定点，循环 repeat_n 次后在测定点采样反馈。"""
    detail: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    awayf = _float_row(away if away is not None else DEFAULT_ANGLE_AWAY)
    if not _in_joint_limits(awayf):
        raise ValueError(f"关节远离点超出限位: {awayf}")

    for pi, cmd in enumerate(targets):
        cmdf = _float_row(cmd)
        if not _in_joint_limits(cmdf):
            raise ValueError(f"关节目标 {pi} 超出限位: {cmdf}")
        if not _poses_differ(awayf, cmdf):
            raise ValueError(f"关节测定点 {pi + 1} 与远离点相同，无法评估重复定位: {cmdf}")
        samples: list[list[float]] = []
        for k in range(repeat_n):
            mc.set_angles(awayf, speed)
            wait_motion_done(mc)
            mc.set_angles(cmdf, speed)
            wait_motion_done(mc)
            act = read_angles_info(mc)
            samples.append(act)
            detail.append(
                {
                    "目标点序号": pi + 1,
                    "重复序号": k + 1,
                    "运动路径": "远离点→测定点",
                    "J1远离点(°)": awayf[0],
                    "J2远离点(°)": awayf[1],
                    "J3远离点(°)": awayf[2],
                    "J4远离点(°)": awayf[3],
                    "J1测定点指令(°)": cmdf[0],
                    "J2测定点指令(°)": cmdf[1],
                    "J3测定点指令(°)": cmdf[2],
                    "J4测定点指令(°)": cmdf[3],
                    "J1反馈(°)": act[0],
                    "J2反馈(°)": act[1],
                    "J3反馈(°)": act[2],
                    "J4反馈(°)": act[3],
                }
            )

        for j, name in enumerate(("J1", "J2", "J3", "J4")):
            col = [s[j] for s in samples]
            row = {
                "目标点序号": pi + 1,
                "关节": name,
                "测定点指令(°)": cmdf[j],
            }
            row.update(_repeat_position_stats(col))
            summary.append(row)

    return detail, summary


def run_absolute_coords(
    mc: UltraArmP1,
    speed: int,
    targets: list[list[float]],
    repeat_n: int,
    away: Sequence[float] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """绝对定位：远离点 → 测定点，循环 repeat_n 次后在测定点采样并统计指令偏差。"""
    detail: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    axis_labels = ("X(mm)", "Y(mm)", "Z(mm)", "Rx(°)")
    away_keys = ("X远离点(mm)", "Y远离点(mm)", "Z远离点(mm)", "Rx远离点(°)")
    awayf = _float_row(away if away is not None else DEFAULT_COORD_AWAY)
    if not _in_coord_limits(awayf):
        raise ValueError(f"笛卡尔远离点超出限位: {awayf}")

    for pi, cmd in enumerate(targets):
        cmdf = _float_row(cmd)
        if not _in_coord_limits(cmdf):
            raise ValueError(f"坐标目标 {pi} 超出限位: {cmdf}")
        if not _poses_differ(awayf, cmdf):
            raise ValueError(f"笛卡尔测定点 {pi + 1} 与远离点相同，无法评估绝对定位: {cmdf}")
        move_to_coords_init(mc, speed)
        errors_by_axis: list[list[float]] = [[], [], [], []]
        for k in range(repeat_n):
            mc.set_coords(awayf, speed)
            wait_motion_done(mc)
            mc.set_coords(cmdf, speed)
            wait_motion_done(mc)
            act = read_coords_info(mc)
            err = [a - c for a, c in zip(act, cmdf)]
            for j in range(4):
                errors_by_axis[j].append(err[j])
            detail.append(
                {
                    "目标点序号": pi + 1,
                    "重复序号": k + 1,
                    "运动路径": "远离点→测定点",
                    away_keys[0]: awayf[0],
                    away_keys[1]: awayf[1],
                    away_keys[2]: awayf[2],
                    away_keys[3]: awayf[3],
                    "X测定点指令(mm)": cmdf[0],
                    "Y测定点指令(mm)": cmdf[1],
                    "Z测定点指令(mm)": cmdf[2],
                    "Rx测定点指令(°)": cmdf[3],
                    "X反馈(mm)": act[0],
                    "Y反馈(mm)": act[1],
                    "Z反馈(mm)": act[2],
                    "Rx反馈(°)": act[3],
                    "X误差(mm)": err[0],
                    "Y误差(mm)": err[1],
                    "Z误差(mm)": err[2],
                    "Rx误差(°)": err[3],
                    "四项最大绝对误差": max(abs(e) for e in err),
                }
            )

        for j, name in enumerate(axis_labels):
            row = {
                "目标点序号": pi + 1,
                "坐标轴": name,
                "测定点指令": cmdf[j],
            }
            row.update(_absolute_error_stats(errors_by_axis[j]))
            summary.append(row)

    return detail, summary


def run_repeat_coords(
    mc: UltraArmP1,
    speed: int,
    targets: list[list[float]],
    repeat_n: int,
    away: Sequence[float] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """重复定位：远离点 → 测定点，循环 repeat_n 次后在测定点采样反馈。"""
    detail: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []

    axis_labels = ("X(mm)", "Y(mm)", "Z(mm)", "Rx(°)")
    away_keys = ("X远离点(mm)", "Y远离点(mm)", "Z远离点(mm)", "Rx远离点(°)")
    awayf = _float_row(away if away is not None else DEFAULT_COORD_AWAY)
    if not _in_coord_limits(awayf):
        raise ValueError(f"笛卡尔远离点超出限位: {awayf}")

    for pi, cmd in enumerate(targets):
        cmdf = _float_row(cmd)
        if not _in_coord_limits(cmdf):
            raise ValueError(f"坐标目标 {pi} 超出限位: {cmdf}")
        if not _poses_differ(awayf, cmdf):
            raise ValueError(f"笛卡尔测定点 {pi + 1} 与远离点相同，无法评估重复定位: {cmdf}")
        move_to_coords_init(mc, speed)
        samples: list[list[float]] = []
        for k in range(repeat_n):
            mc.set_coords(awayf, speed)
            wait_motion_done(mc)
            mc.set_coords(cmdf, speed)
            wait_motion_done(mc)
            act = read_coords_info(mc)
            samples.append(act)
            detail.append(
                {
                    "目标点序号": pi + 1,
                    "重复序号": k + 1,
                    "运动路径": "远离点→测定点",
                    away_keys[0]: awayf[0],
                    away_keys[1]: awayf[1],
                    away_keys[2]: awayf[2],
                    away_keys[3]: awayf[3],
                    "X测定点指令(mm)": cmdf[0],
                    "Y测定点指令(mm)": cmdf[1],
                    "Z测定点指令(mm)": cmdf[2],
                    "Rx测定点指令(°)": cmdf[3],
                    "X反馈(mm)": act[0],
                    "Y反馈(mm)": act[1],
                    "Z反馈(mm)": act[2],
                    "Rx反馈(°)": act[3],
                }
            )

        for j, name in enumerate(axis_labels):
            col = [s[j] for s in samples]
            row = {
                "目标点序号": pi + 1,
                "坐标轴": name,
                "测定点指令": cmdf[j],
            }
            row.update(_repeat_position_stats(col))
            summary.append(row)

    return detail, summary


def _autosize_columns(ws: Any) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        maxlen = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[letter].width = min(maxlen + 2, 48)


def write_workbook(
    path: Path,
    meta: dict[str, Any],
    absolute_angle_detail: list[dict[str, Any]],
    absolute_angle_summary: list[dict[str, Any]],
    repeat_angle_detail: list[dict[str, Any]],
    repeat_angle_summary: list[dict[str, Any]],
    absolute_coord_detail: list[dict[str, Any]],
    absolute_coord_summary: list[dict[str, Any]],
    repeat_coord_detail: list[dict[str, Any]],
    repeat_coord_summary: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "测试说明与参数"
    ws0.append(["项目", "内容"])
    for k, v in meta.items():
        ws0.append([k, str(v)])
    ws0["A1"].font = Font(bold=True)
    ws0["B1"].font = Font(bold=True)
    _autosize_columns(ws0)

    def sheet_from_dicts(name: str, rows: list[dict[str, Any]]) -> None:
        ws = wb.create_sheet(name)
        if not rows:
            ws.append(["本表无数据"])
            return
        headers: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    headers.append(key)
        ws.append(headers)
        for h in ws[1]:
            h.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(k, "") for k in headers])
        _autosize_columns(ws)

    sheet_from_dicts("关节绝对精度_明细", absolute_angle_detail)
    sheet_from_dicts("关节绝对精度_汇总", absolute_angle_summary)
    sheet_from_dicts("关节重复精度_明细", repeat_angle_detail)
    sheet_from_dicts("关节重复精度_汇总", repeat_angle_summary)
    sheet_from_dicts("笛卡尔绝对精度_明细", absolute_coord_detail)
    sheet_from_dicts("笛卡尔绝对精度_汇总", absolute_coord_summary)
    sheet_from_dicts("笛卡尔重复精度_明细", repeat_coord_detail)
    sheet_from_dicts("笛卡尔重复精度_汇总", repeat_coord_summary)

    wb.save(path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="UltraArm P1 定位精度 demo → Excel")
    p.add_argument("--port", default="COM9", help="串口，如 COM10")
    p.add_argument("--baud", type=int, default=1_000_000, help="波特率")
    p.add_argument("--speed", type=int, default=50, help="set_angles / set_coords 速度参数")
    p.add_argument(
        "--repeat",
        type=int,
        default=100,
        help="每个目标点的重复次数（绝对/重复定位精度均适用，默认 100）",
    )
    p.add_argument(
        "--skip-coords",
        action="store_true",
        help="仅测关节空间，跳过笛卡尔段（若逆解不可达可开此项）",
    )
    p.add_argument(
        "--read-retries",
        type=int,
        default=READ_MAX_ATTEMPTS,
        help="get_angles_info/get_coords_info 失败后的最大读取次数（默认 10）",
    )
    p.add_argument(
        "--read-interval",
        type=float,
        default=READ_RETRY_INTERVAL_SEC,
        help="读反馈重试间隔（秒，默认 0.2）",
    )
    p.add_argument(
        "--read-settle",
        type=float,
        default=READ_SETTLE_SEC,
        help="运动结束后首次读反馈前的稳定等待（秒，默认 0.15）",
    )
    p.add_argument(
        "--out",
        type=str,
        default="",
        help="输出 xlsx 路径；默认写入 scripts/P1_positioning_accuracy_<时间戳>.xlsx",
    )
    return p.parse_args()


def main() -> None:
    global READ_MAX_ATTEMPTS, READ_RETRY_INTERVAL_SEC, READ_SETTLE_SEC
    args = parse_args()
    READ_MAX_ATTEMPTS = max(1, int(args.read_retries))
    READ_RETRY_INTERVAL_SEC = max(0.0, float(args.read_interval))
    READ_SETTLE_SEC = max(0.0, float(args.read_settle))
    if args.repeat < 1:
        raise ValueError("--repeat 须 >= 1")
    demo_dir = Path(__file__).resolve().parent
    out_path = (
        Path(args.out)
        if args.out.strip()
        else demo_dir / f"P1_positioning_accuracy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    mc = UltraArmP1(args.port, args.baud, debug=1)
    try:
        angle_targets = [list(map(float, x)) for x in DEFAULT_ANGLE_TARGETS]
        coord_targets = [list(map(float, x)) for x in DEFAULT_COORD_TARGETS]

        absolute_angle_detail, absolute_angle_summary = run_absolute_angles(
            mc, args.speed, angle_targets, args.repeat
        )
        rad, ras = run_repeat_angles(mc, args.speed, angle_targets, args.repeat)

        if args.skip_coords:
            absolute_coord_detail: list[dict[str, Any]] = []
            absolute_coord_summary: list[dict[str, Any]] = []
            rcd: list[dict[str, Any]] = []
            rcs: list[dict[str, Any]] = []
        else:
            absolute_coord_detail, absolute_coord_summary = run_absolute_coords(
                mc, args.speed, coord_targets, args.repeat
            )
            rcd, rcs = run_repeat_coords(mc, args.speed, coord_targets, args.repeat)

        meta = {
            "生成时间": datetime.now().isoformat(timespec="seconds"),
            "串口": args.port,
            "波特率": args.baud,
            "运动速度参数": args.speed,
            "每目标点重复次数": args.repeat,
            "运动路径": "远离点→测定点（绝对/重复定位均在重新逼近后采样）",
            "关节远离点": str(DEFAULT_ANGLE_AWAY),
            "笛卡尔远离点": str(DEFAULT_COORD_AWAY),
            "读反馈重试次数": READ_MAX_ATTEMPTS,
            "读反馈重试间隔(s)": READ_RETRY_INTERVAL_SEC,
            "读反馈稳定等待(s)": READ_SETTLE_SEC,
            "统计说明": "重复定位汇总含均值/方差/标准差/3σ/极差；绝对定位汇总含偏差均值/偏差方差/均方根误差/最大绝对误差（总体方差与标准差）",
            "关节限位J1~J4(°)": str(JOINT_LIMITS_DEG),
            "坐标限位(XYZ/Rx)": str(COORD_LIMITS),
            "备注": "误差来自控制器反馈(get_angles_info/get_coords_info)，非外部激光跟踪仪测量",
            "笛卡尔前置关节角(coords_init)": str(COORDS_INIT_ANGLES),
            "LimitError说明": "固件 LimitError:6 为 J2/J3 耦合限位；笛卡尔段已在每次 set_coords 前回到 coords_init",
            "关节测试指令列表": str(angle_targets),
            "笛卡尔测试指令列表": str(coord_targets),
        }
        write_workbook(
            out_path,
            meta,
            absolute_angle_detail,
            absolute_angle_summary,
            rad,
            ras,
            absolute_coord_detail,
            absolute_coord_summary,
            rcd,
            rcs,
        )
        print(f"已保存: {out_path}")
    finally:
        mc.close()


if __name__ == "__main__":
    main()
