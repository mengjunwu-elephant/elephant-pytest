# -*- coding: utf-8 -*-
"""
将 test_data 下 Excel 规范为单臂（Mercury A1）用例数据。

- 删除 test_type 为 right / exception_right 的行（与左臂镜像的重复用例）。
- 全表字符串单元格：左臂/右臂/左右臂/双臂 → 机械臂（按顺序替换，避免子串问题）。

用法（在项目根目录）::
    python scripts/migrate_test_data_single_arm.py
    python scripts/migrate_test_data_single_arm.py --dry-run

风险边界：若某 sheet 无 test_type 列，仅做文案替换、不删行。
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook

# 先长后短，避免「左右臂」被拆成两次替换
_TEXT_REPLACEMENTS: tuple[tuple[str, str], ...] = (
    ("左右臂", "机械臂"),
    ("左臂", "机械臂"),
    ("右臂", "机械臂"),
    ("双臂", "机械臂"),
)

_TYPES_TO_DROP: frozenset[str] = frozenset({"right", "exception_right"})


def _norm_test_type(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _replace_cell_text(val: object) -> object:
    if not isinstance(val, str) or not val:
        return val
    s = val
    for old, new in _TEXT_REPLACEMENTS:
        s = s.replace(old, new)
    return s


def migrate_workbook(path: Path, *, dry_run: bool) -> dict[str, int]:
    stats = {"sheets_touched": 0, "rows_deleted": 0, "cells_rewritten": 0}
    wb = load_workbook(path, data_only=False)
    try:
        for sheet_name in wb.sheetnames:
            sh = wb[sheet_name]
            # 表头
            headers: list[str] = []
            for c in range(1, sh.max_column + 1):
                v = sh.cell(1, c).value
                headers.append(str(v).strip() if v is not None else "")

            if "test_type" in headers:
                tt_col = headers.index("test_type") + 1
                # 自下而上删行，避免行号漂移
                for r in range(sh.max_row, 1, -1):
                    tt = _norm_test_type(sh.cell(r, tt_col).value)
                    if tt in _TYPES_TO_DROP:
                        sh.delete_rows(r, 1)
                        stats["rows_deleted"] += 1
                stats["sheets_touched"] += 1

            for r in range(1, sh.max_row + 1):
                for c in range(1, sh.max_column + 1):
                    cell = sh.cell(r, c)
                    newv = _replace_cell_text(cell.value)
                    if newv != cell.value:
                        cell.value = newv
                        stats["cells_rewritten"] += 1

        if not dry_run:
            wb.save(path)
    finally:
        wb.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel 单臂数据迁移")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只统计将要删除的行数，不写回文件",
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="指定 xlsx，默认处理 test_data/mercury.xlsx",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    target = args.file or (root / "test_data" / "mercury.xlsx")
    if not target.is_file():
        raise SystemExit(f"文件不存在: {target}")

    if str(target.name).startswith("~$"):
        raise SystemExit("不要对 Excel 临时锁文件操作")

    bak = target.with_suffix(target.suffix + ".bak")
    if not args.dry_run:
        shutil.copy2(target, bak)
        print(f"已备份: {bak}")

    st = migrate_workbook(target, dry_run=args.dry_run)
    print(f"文件: {target}")
    print(f"  涉及含 test_type 的表: {st['sheets_touched']}")
    print(f"  删除行数 (right/exception_right): {st['rows_deleted']}")
    print(f"  改写单元格数: {st['cells_rewritten']}")
    if args.dry_run:
        print("(dry-run，未保存)")


if __name__ == "__main__":
    main()
