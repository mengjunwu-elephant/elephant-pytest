# -*- coding: utf-8 -*-
"""修复 bootstrap 后部分 sheet 表头/列错位问题。"""
from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAIN = os.path.join(BASE, "test_data", "UltraArm_P1.xlsx")
ATT = os.path.join(BASE, "test_data", "UltraArm_P1_Attachments.xlsx")


def _rewrite_sheet(wb, name: str, headers: list, rows: list[list]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def main() -> None:
    shutil.copy2(MAIN, MAIN + ".bak_fix")
    shutil.copy2(ATT, ATT + ".bak_fix")
    wb = load_workbook(MAIN)
    _rewrite_sheet(
        wb,
        "get_run_status",
        ["ID", "title", "api", "parameter", "expect_data", "test_type"],
        [
            [1, "静止时读取运行状态", "get_run_status", None, 0, "normal"],
            [2, "运动中读取运行状态", "get_run_status", None, 1, "normal1"],
        ],
    )
    _rewrite_sheet(
        wb,
        "get_base_io_input",
        ["ID", "title", "api", "pin_no", "expect_data", "test_type"],
        [
            [1, "读取底座IO引脚1", "get_base_io_state", 1, 0, "normal"],
            [2, "引脚号越界", "get_base_io_state", 0, None, "exception"],
        ],
    )
    wb.save(MAIN)

    wb_att = load_workbook(ATT)
    _rewrite_sheet(
        wb_att,
        "close_laser",
        ["ID", "title", "api", "state", "expect_data", "test_type"],
        [[1, "关闭激光", "set_pwm_laser_mode", 0, "ok", "normal"]],
    )
    _rewrite_sheet(
        wb_att,
        "set_gripper_parameter",
        ["ID", "title", "api", "parameter", "value", "expect_data", "test_type"],
        [
            [1, "设置夹爪参数", "set_gripper_parameter", 1, 100, 1, "normal"],
            [2, "地址越界", "set_gripper_parameter", 0, 100, None, "exception"],
        ],
    )
    wb_att.save(ATT)
    print("Fixed Excel sheets.")


if __name__ == "__main__":
    main()
